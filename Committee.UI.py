from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import json, os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
from io import BytesIO

app = Flask(__name__)
app.secret_key = "fast_fintech_bilal_ultimate_v6"
json_file = "committee_data.json"

# --- HELPER FUNCTIONS ---
def load_data():
    if os.path.exists(json_file):
        with open(json_file, "r") as f:
            try:
                data = json.load(f)
                return {int(k): v for k, v in data.items()}
            except: return {}
    return {}

def save_data(data):
    with open(json_file, "w") as f:
        json.dump(data, f, indent=4)

def get_total_collection(data):
    all_payments = sum(m['balance'] for m in data.values())
    done_members = sum(1 for m in data.values() if m.get('committee_status') == 'Done')
    target_per_round = len(data) * 5000
    return all_payments - (done_members * target_per_round)

# --- ROUTES ---
@app.route('/')
def index():
    members = load_data()
    current_collection = get_total_collection(members)
    sorted_members = dict(sorted(members.items(), key=lambda x: x[1]['score'], reverse=True))
    return render_template('index.html', members=sorted_members, total=current_collection)

@app.route('/pay', methods=['POST'])
def pay():
    data = load_data()
    m_id = int(request.form['m_id'])
    mon_num = int(request.form['month'])
    date = int(request.form['date'])
    months = ["Jan", "Feb", "March", "April", "May", "June", "July", "Aug", "Sept", "Oct", "Nov", "Dec"]
    
    if m_id in data and 1 <= mon_num <= 12:
        m_name = months[mon_num-1]
        
        # 1. Duplicate Entry Check
        if m_name in data[m_id]['history']:
            flash(f"Error: {m_name} ki payment pehle se mojood hai!", "danger")
            return redirect(url_for('index'))
        
        # 2. Score & Payment Logic
        p_type = "Late" if date > 5 else "ontime"
        if p_type == "Late":
            data[m_id]['score'] -= 10
            gid = data[m_id]['gurantor_id']
            if gid in data: data[gid]['score'] -= 3
        else: 
            data[m_id]['score'] += 5
            
        data[m_id]['balance'] += 5000
        data[m_id]['history'][m_name] = {"payment_type": p_type}
        
        # 3. Payout Logic
        current_collection = get_total_collection(data)
        target = len(data) * 5000
        if current_collection >= target:
            winner_id = None
            max_score = -9999
            for pid, pdata in data.items():
                if pdata.get('committee_status') != "Done" and pdata['score'] > max_score:
                    max_score = pdata['score']
                    winner_id = pid
            
            if winner_id:
                data[winner_id]['committee_status'] = "Done"
                data[winner_id]['payout_date'] = datetime.now().strftime("%d-%b-%Y")
                flash(f"🎊 MUBARAK! Winner selected: {data[winner_id]['Member_Name']}", "info")

        save_data(data)
        flash(f"Payment Recorded for {m_name}!", "success")
    return redirect(url_for('index'))

@app.route('/manual_payout/<int:m_id>')
def manual_payout(m_id):
    data = load_data()
    if m_id in data:
        data[m_id]['committee_status'] = "Done"
        data[m_id]['payout_date'] = datetime.now().strftime("%d-%b-%Y")
        save_data(data)
        flash(f"Status updated for {data[m_id]['Member_Name']}!", "warning")
    return redirect(url_for('index'))

@app.route('/delete/<int:m_id>/<month>')
def delete_payment(m_id, month):
    data = load_data()
    if m_id in data and month in data[m_id]['history']:
        p_type = data[m_id]['history'][month]['payment_type']
        if p_type == "Late":
            data[m_id]['score'] += 10
            gid = data[m_id]['gurantor_id']
            if gid in data: data[gid]['score'] += 3
        else: data[m_id]['score'] -= 5
        data[m_id]['balance'] -= 5000
        del data[m_id]['history'][month]
        save_data(data)
        flash(f"Record Deleted!", "danger")
    return redirect(url_for('index'))

@app.route('/register', methods=['POST'])
def register():
    data = load_data()
    m_id = int(request.form['id'])
    data[m_id] = {
        'Member_Name': request.form['name'], 'Member_id': m_id, 
        'Member_Gurantor': request.form['g_name'], 'gurantor_id': int(request.form['g_id']), 
        'score': 100, 'balance': 0, 'history': {}, 'committee_status': 'Not Done'
    }
    save_data(data)
    flash("New Member Registered!", "success")
    return redirect(url_for('index'))

@app.route('/reset', methods=['POST'])
def reset_system():
    if request.form.get('confirm', '').upper() == "YES":
        save_data({})
        flash("SYSTEM RESET SUCCESSFUL!", "success")
    else:
        flash("Type YES to confirm!", "danger")
    return redirect(url_for('index'))

@app.route('/export')
def export_pro_report():
    data = load_data()
    if not data:
        flash("Export ke liye data nahi hai!", "danger")
        return redirect(url_for('index'))

    wb = Workbook()
    ws = wb.active
    ws.title = "Current_Committee"

    # --- Excel Professional Styling ---
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Paid
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Unpaid
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Done
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center")

    months_map = ["Jan", "Feb", "March", "April", "May", "June", "July", "Aug", "Sept", "Oct", "Nov", "Dec"]
    headers = ["Rank", "Member Name", "Score", "Total Paid", "Status", "Payout Date"] + months_map + ["Pending", "Risk Alert"]
    
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center_align

    # --- Inserting Data with Pro Logic ---
    sorted_members = sorted(data.values(), key=lambda x: x['score'], reverse=True)
    current_month_idx = datetime.now().month

    for index, d in enumerate(sorted_members, start=2):
        paid_history = d.get("history", {})
        pending_count = 0
        member_balance = d.get('balance', 0)
        status_done = d.get("committee_status", "Not Done")
        payout_dt = d.get("payout_date", "-")

        row_data = [index-1, d['Member_Name'], d['score'], member_balance, status_done, payout_dt]
        
        # Monthly statuses
        for i in range(12):
            m_name = months_map[i]
            if m_name in paid_history:
                row_data.append("PAID")
            elif i < current_month_idx:
                row_data.append("UNPAID")
                pending_count += 1
            else:
                row_data.append("-")
        
        # Risk Alert Logic
        alert = "Safe"
        if 1 <= pending_count <= 2: alert = f"Pending: {pending_count}"
        elif 3 <= pending_count <= 4: alert = "WARNING: Guarantor Alert!"
        elif pending_count >= 5: alert = "TERMINATED"
        row_data.extend([pending_count, alert])

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=index, column=col, value=val)
            cell.alignment = center_align
            if val == "PAID": cell.fill = green_fill
            elif val == "UNPAID": cell.fill = red_fill
            if status_done == "Done" and col == 5: cell.fill = gold_fill

    # --- Formatting Widths ---
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['T'].width = 25 # Risk alert column

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, as_attachment=True, download_name=f"Committee_Master_{datetime.now().strftime('%d_%b')}.xlsx")

if __name__ == '__main__':
    app.run(debug=True)